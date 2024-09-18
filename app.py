import tkinter as tk
from tkinter import filedialog, messagebox
from tkinter.ttk import Progressbar
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import webbrowser
import threading
import time

def process_csv():
    try:
        file_path = filedialog.askopenfilename(filetypes=[("CSV files", "*.csv")])
        if file_path:
            progress_bar["value"] = 0
            progress_label.config(text="Processing: 0%")
            window.update_idletasks()

            df = pd.read_csv(file_path, low_memory=False)
            columns_to_delete = ['Discom', 'Zone', 'Substation Name(Master)', 'Substation Code(Master)',
                                'Feeder Name(Master)', 'Feeder Code(Master)', 'DTR Name(Master)', 'Alternate MobileNo',
                                'Email ID', 'Old meter Current Kva CI', 'Old meter Current Kvah CI',
                                'Number of Joints In Incoming Service Cable', 'Account No', 'No Of Rejections']
            df.drop(columns=columns_to_delete, axis=1, inplace=True)

            new_order = ['Survey done by', 'CI Date', 'Sub Division', 'Status', 'QC Contractor Rejected Remarks', 
                         'Sl No', 'Sequence No', 'Circle', 'Division', 'DTR Code(Master)', 'Substation Name', 
                         'Substation Code', 'Feeder Name', 'Feeder Code', 'DTR Name', 'DTR Code', 'Consumer Number', 
                         'Landmark', 'Consumer Name', 'Address', 'Address from field', 'Area', 'Mobile Number', 
                         'Billing Type', 'Old Meter Number', 'Meter Number from Field', 'Mismatch between master and Meter Number from field',
                         'Category Code', 'Old Meter Manufacturer', 'Clear Line of Sight', 'Contracted load', 'Load Unit',
                         'Existing Meter seal Status', 'Existing Meter Available Status', 'JE Name', 
                         'Meter Box Sealing Status', 'Meter Installed with Metallic Enclose', 'Meter Location', 
                         'Meter Shifting required', 'Old Meter Status', 'Old meter Current Kwh CI', 'Mismatch between master and CI reading',
                         'Old Meter MF ', 'Old Meter Phase', 'Service Cable Type', 'Service Line Status', 'Status of Incoming  Service cable ', 
                         'Sub Category Code', 'Town Name', 'Network Provider Name SIM 1', 'Signal Strength SIM 1', 'Signal Level SIM 1', 
                         'Signal Category SIM 1', 'Network Provider Name SIM 2', 'Signal Strength SIM 2', 'Signal Level SIM 2', 
                         'Signal Category SIM 2', 'Latitude', 'Longitude', 'Type', 'Contractor Name', 'CI Remarks', 
                         'QC Contractor Approved Date', 'Approved By QC Contractor', 'IntelliSmart Approved Date', 
                         'Approved By IntelliSmart', 'Discom Approved Date', 'Approved By Discom', 'QC Contractor Rejected Date', 
                         'Rejected By QC Contractor', 'IntelliSmart Rejected Remarks', 'IntelliSmart Rejected Date', 
                         'Rejected By IntelliSmart', 'Discom Rejected Remarks', 'Discom Rejected Date', 'Rejected By Discom']

            df = df[new_order]
            df = df[df['Status'] == 'Pending@Resurvey']
            df = df.sort_values(by='Survey done by')

            global df_sdo1, df_sdo2
            df_sdo1 = df[df['Sub Division'] == 'EUDSD I NOIDA-SDO1410111']
            df_sdo2 = df[df['Sub Division'] == 'EUDSD-4 NOIDA-SDO1410114']

            progress_bar["value"] = 100
            progress_label.config(text="Processing: 100% Complete")
            get_report_button.config(state="normal")

    except Exception as e:
        messagebox.showerror("Error", f"An error occurred: {str(e)}")

def style_excel(file_path):
    try:
        wb = load_workbook(file_path)
        ws = wb.active
        header_fill = PatternFill(start_color='ADD8E6', end_color='ADD8E6', fill_type='solid')
        for cell in ws[1]:
            cell.fill = header_fill
        for col in ws.columns:
            max_length = max(len(str(cell.value)) for cell in col)
            adjusted_width = max_length + 2
            ws.column_dimensions[col[0].column_letter].width = adjusted_width
        wb.save(file_path)
    except Exception as e:
        messagebox.showerror("Error", f"Error styling the file: {str(e)}")

def save_reports():
    try:
        save_path = filedialog.askdirectory()
        if save_path:
            sdo1_file = f"{save_path}/Sub_Div-1_Resurvey.xlsx"
            sdo2_file = f"{save_path}/Sub_Div-2_Resurvey.xlsx"
            
            # Simulate saving time for countdown effect
            for i in range(10, 0, -1):
                countdown_label.config(text=f"Saving in {i} seconds...")
                window.update_idletasks()
                time.sleep(1)

            df_sdo1.to_excel(sdo1_file, index=False)
            df_sdo2.to_excel(sdo2_file, index=False)

            # Run styling in separate threads
            threading.Thread(target=style_excel, args=(sdo1_file,)).start()
            threading.Thread(target=style_excel, args=(sdo2_file,)).start()

            countdown_label.config(text="Reports saved successfully!")
            messagebox.showinfo("Success", f"Reports saved at: {save_path}")
    except Exception as e:
        messagebox.showerror("Error", f"Error saving reports: {str(e)}")

def get_report():
    threading.Thread(target=save_reports).start()

def open_linkedin():
    webbrowser.open("https://www.linkedin.com/in/soumy-chauhan/")

def open_instagram():
    webbrowser.open("https://www.instagram.com/mrx_3.2/")

def create_gradient(canvas, color1, color2):
    width = canvas.winfo_width()
    height = canvas.winfo_height()
    for i in range(height):
        color = "#%02x%02x%02x" % (
            int(color1[0] + (color2[0] - color1[0]) * i / height),
            int(color1[1] + (color2[1] - color1[1]) * i / height),
            int(color1[2] + (color2[2] - color1[2]) * i / height)
        )
        canvas.create_line(0, i, width, i, fill=color, width=1)

window = tk.Tk()
window.title("mrx_3.2")
window.geometry("550x480")
window.configure(bg="#f0f0f0")

# Create canvas for gradient background
canvas = tk.Canvas(window, width=550, height=480)
canvas.pack(fill="both", expand=True)

# Create gradient effect on canvas
create_gradient(canvas, (230, 240, 255), (200, 220, 255))

# Create frame to hold widgets
frame = tk.Frame(canvas, bg="#f3f0f0")
frame.place(relwidth=1, relheight=1)

title_label = tk.Label(frame, text="mrx_3.2", font=("Arial", 18, "bold"), bg="#f0f0f0", fg="#333")
title_label.pack(pady=10)

upload_button = tk.Button(frame, text="Upload CI_Report CSV", command=process_csv, font=("Arial", 12), bg="#007acc", fg="white", padx=20, pady=10, bd=0)
upload_button.pack(pady=20)

progress_bar = Progressbar(frame, orient="horizontal", length=400, mode="determinate")
progress_bar.pack(pady=10)

progress_label = tk.Label(frame, text="Processing: 0%", font=("Arial", 10), bg="#f0f0f0", fg="#333")
progress_label.pack()

countdown_label = tk.Label(frame, text="", font=("Arial", 10), bg="#f0f0f0", fg="#333")
countdown_label.pack(pady=10)

get_report_button = tk.Button(frame, text="Get Reports", command=get_report, font=("Arial", 12), bg="#28a745", fg="white", padx=20, pady=10, bd=0, state="disabled")
get_report_button.pack(pady=5)

developer_frame = tk.Frame(frame, bg="#f0f0f0")
developer_frame.pack(side="bottom", pady=20)

dev_label = tk.Label(developer_frame, text="Developed by Soumy Chauhan", font=("Arial", 10), bg="#f0f0f0", fg="#333")
dev_label.pack()

linkedin_button = tk.Button(developer_frame, text="LinkedIn", font=("Arial", 10), bg="#0077b5", fg="white", command=open_linkedin, bd=0, padx=10, pady=5)
linkedin_button.pack(side="left", padx=10)

instagram_button = tk.Button(developer_frame, text="Instagram", font=("Arial", 10), bg="#bc2a8d", fg="white", command=open_instagram, bd=0, padx=10, pady=5)
instagram_button.pack(side="right", padx=10)

footer_label = tk.Label(frame, text="All Rights Reserved © 2024 Soumy Chauhan", font=("Arial", 8), bg="#f0f2f0", fg="#777")
footer_label.pack(side="bottom", pady=10)

def on_hover(button, color_on_hover, color_on_leave):
    button.bind("<Enter>", lambda e: button.config(bg=color_on_hover))
    button.bind("<Leave>", lambda e: button.config(bg=color_on_leave))

on_hover(upload_button, "#005f99", "#007acc")
on_hover(get_report_button, "#1e7b34", "#28a745")
on_hover(linkedin_button, "#005582", "#0077b5")
on_hover(instagram_button, "#99205f", "#bc2a8d")

window.mainloop()
